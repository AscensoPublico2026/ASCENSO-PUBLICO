# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1,<4"]
# ///
"""Genera el índice web unificado de vacantes de Ascenso Público.

Integra tres fuentes:
- Empresas Sociales del Estado 2026 (ESE 2).
- Entidades del Orden Nacional 2026 (EREON).
- Procuraduría General de la Nación 2026.

Uso desde plataforma/:
    uv run scripts/generar_datos_opec.py --fecha 2026-07-15

Cuando un Excel nuevo aún no está materializado en el checkout, el script usa su
archivo equivalente en la rama main. Para ESE conserva compatibilidad con los
fragmentos JSON ya generados, de modo que no se pierda el buscador existente.
"""

from __future__ import annotations

import argparse
import html
import io
import json
import re
import unicodedata
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
INDEX_FILE = REPO_ROOT / "plataforma" / "public" / "data" / "opec-index.json"
META_FILE = REPO_ROOT / "plataforma" / "data" / "opec-meta.json"
DETAILS_DIR = REPO_ROOT / "plataforma" / "public" / "data" / "opec-details"
DETAIL_SHARDS = 32
RAW_BASE_URL = "https://raw.githubusercontent.com/AscensoPublico2026/ASCENSO-PUBLICO/main"

ESE_SOURCE = REPO_ROOT / "Empleos_SIMO_ESEs_2026_clae-2.xlsx"
EREON_SOURCE = REPO_ROOT / "Empleos_SIMO_EREON2026_clave-2-2.xlsx"
PROCURADURIA_SOURCE = REPO_ROOT / "Convocatorias_Procuraduria.xlsx"

CONVOCATORIAS = {
    "ese-2026": "Empresas Sociales del Estado 2026",
    "ereon-2026": "Entidades del Orden Nacional 2026",
    "procuraduria-2026": "Procuraduría General de la Nación 2026",
}

# Dato confirmado directamente por Ascenso Público el 15 de julio de 2026.
CORRECCIONES_ESE = {245015: {"Total vacantes": 5, "Disponibles": 5}}

PROCESS_LABELS = {
    "CONCURSO_ABIERTO": "Abierto",
    "CONCURSO_ABIERTO_DISCAPACIDAD": "Abierto",
    "CONCURSO_CERRADO_ASCENSOS": "Ascenso",
    "CONCURSO_CERRADO_ASCENSOS_DISCAPACIDAD": "Ascenso",
}

STUDY_LABELS = {
    "primaria-laboral": "Primaria + formación laboral",
    "secundaria-aprobada": "4 o 5 años de bachillerato aprobados",
    "bachiller-solo": "Solo título de bachiller",
    "bachiller-curso": "Bachiller + curso o certificación",
    "tecnico-laboral": "Técnico laboral o auxiliar",
    "tecnico-tecnologo": "Técnico profesional o tecnólogo",
    "universitario-sin-titulo": "Estudios universitarios sin título",
    "profesional": "Título profesional universitario",
    "posgrado": "Título profesional + posgrado",
    "curso-especifico": "Curso o certificación específica",
}

STOP_WORDS = {
    "para", "como", "esta", "este", "estos", "estas", "desde", "hasta", "entre",
    "sobre", "segun", "tiene", "tener", "debe", "deben", "donde", "cuando", "cada",
    "todo", "toda", "todos", "todas", "mismo", "misma", "demas", "acuerdo", "cargo",
    "empleo", "entidad", "funciones", "funcion", "actividades", "actividad", "proceso",
    "procesos", "procedimiento", "procedimientos", "realizar", "ejecutar", "apoyar",
    "participar", "cumplir", "garantizar", "brindar", "nivel", "area", "servicio",
    "servicios", "informacion", "requerida", "requerido", "correspondiente", "vigente",
    "normas", "normatividad", "lineamientos", "establecidos", "establecidas", "manera",
    "mediante", "conforme", "institucional", "institucionales", "desarrollo", "gestion",
    "profesional", "experiencia", "titulo", "formacion", "meses", "anos", "trabajo",
    "relacionada", "relacionado", "general", "publica", "publico", "persona", "personas",
    "otras", "otros", "podra", "sean", "asignadas", "asignados", "competencia",
}


def texto(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def limpiar_html(value: Any) -> str:
    raw = texto(value)
    raw = re.sub(r"<br\s*/?>", " ", raw, flags=re.IGNORECASE)
    raw = re.sub(r"<[^>]+>", "", raw)
    return re.sub(r"\s+", " ", html.unescape(raw)).strip(" ,")


def normalizar_busqueda(value: Any) -> str:
    normalized = unicodedata.normalize("NFD", limpiar_html(value).lower())
    normalized = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", normalized).strip()


def numero_entero(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, str):
        value = value.replace("$", "").replace(".", "").replace(",", "").strip()
    return int(float(value))


def clasificar_estudio(value: Any) -> tuple[str, str]:
    requirement = normalizar_busqueda(value).upper()
    has_primary = "PRIMARIA" in requirement
    has_bachelor = "BACHILLERATO" in requirement or bool(re.search(r"\bBACHILLER\b", requirement))
    has_job_training = any(
        marker in requirement
        for marker in ("EDUCACION PARA EL TRABAJO", "FORMACION LABORAL", "TECNICO LABORAL", "AUXILIAR DE")
    )
    has_technical_professional = any(
        marker in requirement
        for marker in ("TECNICA PROFESIONAL", "FORMACION TECNICA PROFESIONAL", "TECNOLOGICA", "TECNOLOGIA EN")
    )
    has_professional_degree = bool(
        re.search(r"TITULO (?:DE FORMACION UNIVERSITARIA|DE PROFESIONAL|PROFESIONAL EN)", requirement)
    ) or "FORMACION UNIVERSITARIA EN" in requirement
    has_partial_university = bool(
        re.search(r"(?:APROBACION|TERMINACION).*?(?:ANOS|SEMESTRES|PENSUM).*?PROFESIONAL", requirement)
    )
    has_postgraduate = any(
        marker in requirement
        for marker in ("POSTGRADO", "ESPECIALIZACION", "MAESTRIA", "DOCTORADO")
    )
    bachelor_only = bool(re.fullmatch(r"(?:FORMACION BASICA )?TITULO DE BACHILLERATO\.?", requirement))
    has_partial_secondary = bool(re.search(r"APROBACION DE .*?ANOS.*?BACHILLERATO", requirement))

    if has_primary:
        category = "primaria-laboral"
    elif has_postgraduate:
        category = "posgrado"
    elif has_job_training:
        category = "tecnico-laboral"
    elif has_technical_professional:
        category = "tecnico-tecnologo"
    elif has_professional_degree:
        category = "profesional"
    elif has_partial_university:
        category = "universitario-sin-titulo"
    elif has_partial_secondary:
        category = "secundaria-aprobada"
    elif bachelor_only:
        category = "bachiller-solo"
    elif has_bachelor:
        category = "bachiller-curso"
    else:
        category = "curso-especifico"

    return category, STUDY_LABELS[category]


def resumir_requisito(value: Any, max_length: int = 170) -> str:
    requirement = limpiar_html(value)
    if len(requirement) <= max_length:
        return requirement
    shortened = requirement[: max_length - 1].rsplit(" ", 1)[0]
    return f"{shortened}…"


def clasificar_requisitos_adicionales(value: Any) -> tuple[list[str], list[str]]:
    requirement = normalizar_busqueda(value).upper()
    if not requirement:
        return [], []

    keys: list[str] = []
    labels: list[str] = []
    rules = (
        ("rethus", "ReTHUS", ("RETHUS", "TALENTO HUMANO EN SALUD")),
        ("tarjeta-profesional", "Tarjeta o matrícula profesional", ("TARJETA PROFESIONAL", "TARJETA O MATRICULA PROFESIONAL", "MATRICULA PROFESIONAL")),
        ("licencia-conduccion", "Licencia de conducción", ("LICENCIA DE CONDUCCION",)),
        ("licencia-sst", "Licencia en seguridad y salud en el trabajo", ("LICENCIA VIGENTE EN SEGURIDAD Y SALUD",)),
        ("registro-archivistas", "Registro profesional de archivistas", ("REGISTRO UNICO PROFESIONAL DE ARCHIVISTAS",)),
        ("soporte-vital", "Soporte vital básico", ("SOPORTE VITAL BASICO",)),
        ("radioproteccion", "Carné de radioprotección", ("RADIO PROTECCION", "RADIOPROTECCION")),
        ("antecedentes-contador", "Antecedentes de contador", ("ANTECEDENTES DE CONTADOR",)),
        ("cedula", "Cédula de ciudadanía", ("CEDULA DE CIUDADANIA",)),
    )
    for key, label, markers in rules:
        if any(marker in requirement for marker in markers):
            keys.append(key)
            labels.append(label)

    if not labels:
        return ["otro"], ["Otro requisito adicional"]
    return keys, labels


def experiencia_meses(value: Any) -> int | None:
    requirement = limpiar_html(value)
    if not requirement:
        return None
    if re.search(r"\b(?:no requiere|sin)\s+experiencia\b", requirement, re.IGNORECASE):
        return 0

    amounts: list[int] = []
    for amount, unit in re.findall(
        r"\((\d+)\)\s*(año(?:s)?|mes(?:es)?)\b", requirement, flags=re.IGNORECASE
    ):
        amounts.append(int(amount) * (12 if unit.lower().startswith("año") else 1))

    if not amounts:
        for amount, unit in re.findall(r"\b(\d+)\s*(año(?:s)?|mes(?:es)?)\b", requirement, re.IGNORECASE):
            amounts.append(int(amount) * (12 if unit.lower().startswith("año") else 1))
    if not amounts:
        return None

    has_alternative = bool(re.search(r"(?:^|[,;\s])O(?:[,;\s]|$)", requirement, re.IGNORECASE))
    return min(amounts) if has_alternative else max(amounts)


def lista_funciones(value: Any) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    if "||" in raw:
        chunks = raw.split("||")
    elif re.search(r"(?:^|\s)\d+[.)]\s+", raw):
        chunks = re.split(r"(?=(?:^|\s)\d+[.)]\s+)", raw)
    else:
        chunks = raw.splitlines()
    return [limpiar_html(item) for item in chunks if limpiar_html(item)]


def valores_multiples(value: Any, empty_label: str = "No especificado") -> str:
    raw = texto(value).replace(";", "|")
    items = [item.strip().replace("No_Aplica", "No aplica") for item in raw.split("|") if item.strip()]
    return " | ".join(dict.fromkeys(items)) if items else empty_label


def palabras_clave(weighted_texts: Iterable[tuple[str, int]], limit: int = 90) -> str:
    frequencies: Counter[str] = Counter()
    for value, weight in weighted_texts:
        tokens = normalizar_busqueda(value).split()
        frequencies.update(
            token
            for token in tokens
            for _ in range(weight)
            if len(token) >= 4 and token not in STOP_WORDS and not token.isdigit()
        )
    return " ".join(token for token, _ in frequencies.most_common(limit))


def referencia_id(convocatoria_id: str, reference: str) -> str:
    return f"{convocatoria_id}:{normalizar_busqueda(reference).replace(' ', '-')}"


def construir_registro(
    *,
    convocatoria_id: str,
    reference: Any,
    tipo_referencia: str,
    denomination: Any,
    level: Any,
    grade: Any,
    code: Any,
    salary: Any,
    salary_year: Any,
    entity: Any,
    department: Any,
    municipality: Any,
    vacancies: Any,
    available: Any,
    modality: str,
    disability: bool,
    study: Any,
    experience: Any,
    other_requirement: Any,
    purpose: Any,
    functions: Any,
    dependency: Any,
    nit: Any = "",
    entity_type: Any = "",
    detail_call: Any = "",
    year: Any = 2026,
    process_type: Any = "",
    promotion: bool = False,
    official_url: str = "",
) -> tuple[dict[str, Any], dict[str, Any]]:
    reference_text = texto(reference)
    identifier = referencia_id(convocatoria_id, reference_text)
    study_text = limpiar_html(study)
    experience_text = limpiar_html(experience)
    other_text = limpiar_html(other_requirement)
    purpose_text = limpiar_html(purpose)
    function_list = lista_funciones(functions) if not isinstance(functions, list) else [limpiar_html(item) for item in functions if limpiar_html(item)]
    dependency_text = texto(dependency)
    study_level, study_summary = clasificar_estudio(study_text)
    additional_keys, additional_labels = clasificar_requisitos_adicionales(other_text)
    months = experiencia_meses(experience_text)
    denomination_text = texto(denomination)
    entity_text = texto(entity)
    department_text = valores_multiples(department)
    municipality_text = valores_multiples(municipality)
    level_text = texto(level)
    profile_functions = palabras_clave(
        [
            (denomination_text, 6),
            (dependency_text, 3),
            (purpose_text, 3),
            (" ".join(function_list), 1),
        ]
    )
    profile_education = normalizar_busqueda(study_text)
    profile_experience = normalizar_busqueda(experience_text)
    search_text = normalizar_busqueda(
        " ".join(
            [
                reference_text,
                tipo_referencia,
                CONVOCATORIAS[convocatoria_id],
                denomination_text,
                entity_text,
                department_text,
                municipality_text,
                level_text,
                texto(code),
                dependency_text,
                study_text,
                experience_text,
                purpose_text,
                profile_functions,
                "sin experiencia" if months == 0 else "",
            ]
        )
    )

    summary = {
        "id": identifier,
        "referencia": reference_text,
        "tipoReferencia": tipo_referencia,
        "convocatoriaId": convocatoria_id,
        "convocatoriaNombre": CONVOCATORIAS[convocatoria_id],
        "denominacion": denomination_text,
        "nivel": level_text,
        "grado": texto(grade),
        "codigo": texto(code),
        "salario": numero_entero(salary),
        "vigenciaSalario": numero_entero(salary_year),
        "entidad": entity_text,
        "departamento": department_text,
        "municipio": municipality_text,
        "vacantes": numero_entero(vacancies),
        "disponibles": numero_entero(available),
        "modalidad": modality,
        "discapacidad": disability,
        "sinExperiencia": months == 0,
        "experienciaMeses": months,
        "nivelEstudio": study_level,
        "estudioResumen": study_summary,
        "estudioDetalleCorto": resumir_requisito(study_text),
        "requisitosAdicionales": additional_keys,
        "requisitosAdicionalesResumen": additional_labels,
        "perfilFormacion": profile_education,
        "perfilExperiencia": profile_experience,
        "perfilFunciones": profile_functions,
        "busqueda": search_text,
    }
    detail = {
        **{
            key: value
            for key, value in summary.items()
            if key not in {"busqueda", "perfilFormacion", "perfilExperiencia", "perfilFunciones"}
        },
        "nit": texto(nit),
        "tipoEntidad": texto(entity_type),
        "convocatoriaDetalle": texto(detail_call) or CONVOCATORIAS[convocatoria_id],
        "anio": numero_entero(year),
        "tipoProceso": texto(process_type),
        "dependencia": dependency_text,
        "concursoAscenso": promotion,
        "proposito": purpose_text,
        "funciones": function_list,
        "requisitoEstudio": study_text,
        "requisitoExperiencia": experience_text,
        "requisitoOtros": other_text,
        "urlOficial": official_url,
    }
    return summary, detail


def abrir_excel(path: Path):
    if path.exists():
        print(f"Fuente local: {path.name}")
        return load_workbook(path, read_only=True, data_only=True)
    url = f"{RAW_BASE_URL}/{urllib.parse.quote(path.name)}"
    print(f"Fuente remota: {url}")
    request = urllib.request.Request(url, headers={"User-Agent": "Ascenso-Publico-data-generator"})
    with urllib.request.urlopen(request, timeout=90) as response:
        return load_workbook(io.BytesIO(response.read()), read_only=True, data_only=True)


def iterar_filas(workbook, sheet_name: str) -> Iterable[dict[str, Any]]:
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [texto(value) for value in next(rows)]
    for values in rows:
        row = dict(zip(headers, values))
        if any(value not in (None, "") for value in values):
            yield row


def cargar_simo(path: Path, convocatoria_id: str) -> list[tuple[dict[str, Any], dict[str, Any]]]:
    workbook = abrir_excel(path)
    records: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for row in iterar_filas(workbook, "Empleos SIMO"):
        opec = numero_entero(row.get("OPEC"))
        if not opec:
            continue
        if convocatoria_id == "ese-2026":
            for field, corrected_value in CORRECCIONES_ESE.get(opec, {}).items():
                row[field] = corrected_value
        process = texto(row.get("Tipo proceso"))
        modality = PROCESS_LABELS.get(process)
        if modality is None:
            raise ValueError(f"Tipo de proceso desconocido en OPEC {opec}: {process}")
        records.append(
            construir_registro(
                convocatoria_id=convocatoria_id,
                reference=opec,
                tipo_referencia="OPEC",
                denomination=row.get("Denominación"),
                level=row.get("Nivel"),
                grade=row.get("Grado"),
                code=row.get("Código empleo"),
                salary=row.get("Asignación salarial"),
                salary_year=row.get("Vigencia salarial"),
                entity=row.get("Entidad"),
                department=row.get("Departamento"),
                municipality=row.get("Municipio"),
                vacancies=row.get("Total vacantes"),
                available=row.get("Disponibles"),
                modality=modality,
                disability="DISCAPACIDAD" in process,
                study=row.get("Requisito estudio"),
                experience=row.get("Requisito experiencia"),
                other_requirement=row.get("Requisito otros"),
                purpose=row.get("Propósito"),
                functions=row.get("Funciones"),
                dependency=row.get("Dependencia"),
                nit=row.get("NIT"),
                entity_type=row.get("Tipo entidad"),
                detail_call=row.get("Convocatoria"),
                year=row.get("Año"),
                process_type=process,
                promotion=texto(row.get("Concurso ascenso")).lower() in {"sí", "si"},
                official_url="https://simo.cnsc.gov.co/",
            )
        )
    workbook.close()
    return records


def cargar_procuraduria() -> list[tuple[dict[str, Any], dict[str, Any]]]:
    workbook = abrir_excel(PROCURADURIA_SOURCE)
    records: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for row in iterar_filas(workbook, "Convocatorias"):
        code = texto(row.get("Código"))
        if not code:
            continue
        records.append(
            construir_registro(
                convocatoria_id="procuraduria-2026",
                reference=code,
                tipo_referencia="Código",
                denomination=row.get("Cargo"),
                level=row.get("Nivel"),
                grade=row.get("Grado"),
                code=row.get("Grado"),
                salary=row.get("Salario mensual"),
                salary_year=2026,
                entity="Procuraduría General de la Nación",
                department=row.get("Departamentos"),
                municipality=row.get("Municipios"),
                vacancies=row.get("N° de plazas"),
                available=row.get("N° de plazas"),
                modality="Abierto",
                disability=False,
                study=row.get("Requisito de estudio"),
                experience=row.get("Requisito de experiencia"),
                other_requirement=row.get("Otros requisitos"),
                purpose=row.get("Propósito del cargo"),
                functions=row.get("Funciones"),
                dependency=row.get("Dependencia"),
                detail_call="Convocatoria Procuraduría General de la Nación 2026",
                year=2026,
                process_type=texto(row.get("Proceso")) or "Concurso abierto",
                promotion=False,
            )
        )
    workbook.close()
    return records


def cargar_ese_legado() -> list[tuple[dict[str, Any], dict[str, Any]]]:
    records: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for shard_path in sorted(DETAILS_DIR.glob("*.json")):
        payload = json.loads(shard_path.read_text(encoding="utf-8"))
        for old in payload.get("empleos", {}).values():
            old_group = old.get("convocatoriaId")
            if old_group not in (None, "ese-2026"):
                continue
            reference = old.get("referencia", old.get("opec"))
            if reference in (None, ""):
                continue
            records.append(
                construir_registro(
                    convocatoria_id="ese-2026",
                    reference=reference,
                    tipo_referencia="OPEC",
                    denomination=old.get("denominacion"),
                    level=old.get("nivel"),
                    grade=old.get("grado"),
                    code=old.get("codigo"),
                    salary=old.get("salario"),
                    salary_year=old.get("vigenciaSalario"),
                    entity=old.get("entidad"),
                    department=old.get("departamento"),
                    municipality=old.get("municipio"),
                    vacancies=old.get("vacantes"),
                    available=old.get("disponibles"),
                    modality=old.get("modalidad", "Abierto"),
                    disability=bool(old.get("discapacidad")),
                    study=old.get("requisitoEstudio"),
                    experience=old.get("requisitoExperiencia"),
                    other_requirement=old.get("requisitoOtros"),
                    purpose=old.get("proposito"),
                    functions=old.get("funciones", []),
                    dependency=old.get("dependencia"),
                    nit=old.get("nit"),
                    entity_type=old.get("tipoEntidad"),
                    detail_call=old.get("convocatoriaDetalle", old.get("convocatoria")),
                    year=old.get("anio", 2026),
                    process_type=old.get("tipoProceso"),
                    promotion=bool(old.get("concursoAscenso")),
                    official_url=old.get("urlOficial", "https://simo.cnsc.gov.co/"),
                )
            )
    if not records:
        raise FileNotFoundError("No se encontró el Excel ESE ni un consolidado JSON legado")
    print(f"Fuente compatible ESE: {len(records)} detalles existentes")
    return records


def split_values(value: str) -> set[str]:
    return {item.strip() for item in value.split("|") if item.strip() and item.strip() != "No especificado"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--fecha", default=date.today().isoformat(), help="Fecha pública AAAA-MM-DD")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ese_records = cargar_simo(ESE_SOURCE, "ese-2026") if ESE_SOURCE.exists() else cargar_ese_legado()
    records = ese_records + cargar_simo(EREON_SOURCE, "ereon-2026") + cargar_procuraduria()

    summaries: list[dict[str, Any]] = []
    details: dict[str, dict[str, Any]] = {}
    for summary, detail in records:
        identifier = summary["id"]
        if identifier in details:
            raise ValueError(f"Identificador duplicado: {identifier}")
        if summary["vacantes"] < 0:
            raise ValueError(f"Vacante con número de plazas negativo: {identifier}")
        summaries.append(summary)
        details[identifier] = detail

    summaries.sort(key=lambda item: (item["convocatoriaNombre"], item["referencia"]))
    groups = []
    for convocatoria_id, name in CONVOCATORIAS.items():
        group_jobs = [item for item in summaries if item["convocatoriaId"] == convocatoria_id]
        groups.append(
            {
                "id": convocatoria_id,
                "nombre": name,
                "totalOpec": len(group_jobs),
                "totalVacantes": sum(item["vacantes"] for item in group_jobs),
            }
        )

    departments: set[str] = set()
    municipalities: set[str] = set()
    for item in summaries:
        departments.update(split_values(item["departamento"]))
        municipalities.update(split_values(item["municipio"]))
    open_jobs = [item for item in summaries if item["modalidad"] == "Abierto"]
    meta = {
        "convocatoria": "Buscador unificado de convocatorias 2026",
        "fechaActualizacion": args.fecha,
        "totalOpec": len(summaries),
        "totalVacantes": sum(item["vacantes"] for item in summaries),
        "opecAbiertas": len(open_jobs),
        "vacantesAbiertas": sum(item["vacantes"] for item in open_jobs),
        "entidades": len({item["entidad"] for item in summaries}),
        "departamentos": len(departments),
        "municipios": len(municipalities),
        "sinExperiencia": sum(item["sinExperiencia"] for item in open_jobs),
        "totalConvocatorias": len(groups),
        "convocatorias": groups,
        "fuente": "Consolidados oficiales suministrados para ESE 2, EREON y Procuraduría",
    }

    if len(summaries) != len(details):
        raise ValueError("El índice y los detalles no tienen la misma cantidad de vacantes")
    expected_groups = {"ese-2026": 1187, "ereon-2026": 707, "procuraduria-2026": 296}
    actual_groups = {group["id"]: group["totalOpec"] for group in groups}
    if actual_groups != expected_groups:
        raise ValueError(f"Conteos inesperados por convocatoria: {actual_groups}")

    INDEX_FILE.parent.mkdir(parents=True, exist_ok=True)
    META_FILE.parent.mkdir(parents=True, exist_ok=True)
    DETAILS_DIR.mkdir(parents=True, exist_ok=True)
    for old_shard in DETAILS_DIR.glob("*.json"):
        old_shard.unlink()

    INDEX_FILE.write_text(
        json.dumps({"meta": meta, "empleos": summaries}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    META_FILE.write_text(json.dumps(meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    shards: list[dict[str, dict[str, Any]]] = [dict() for _ in range(DETAIL_SHARDS)]
    for identifier, detail in details.items():
        shard_number = sum(ord(character) for character in identifier) % DETAIL_SHARDS
        shards[shard_number][identifier] = detail
    for shard_number, shard in enumerate(shards):
        shard_file = DETAILS_DIR / f"{shard_number:02d}.json"
        shard_file.write_text(
            json.dumps({"empleos": shard}, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )

    detail_size = sum(path.stat().st_size for path in DETAILS_DIR.glob("*.json"))
    print(f"Índice: {INDEX_FILE} ({INDEX_FILE.stat().st_size / 1024:.1f} KiB)")
    print(f"Detalles: {DETAIL_SHARDS} fragmentos ({detail_size / 1024:.1f} KiB)")
    print(json.dumps(meta, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
